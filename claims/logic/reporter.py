"""
Summary and CT600L Output Logic
Generates reports and exports for R&D claims
"""

import pandas as pd
from typing import Dict, List, Optional
from decimal import Decimal
from datetime import datetime, date
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


class RDReporter:
    """Generates reports and exports for R&D claims"""
    
    def __init__(self, processed_results: Dict, claim_info: Dict):
        """
        Initialize the reporter with processed results and claim information
        
        Args:
            processed_results: Results from RDProcessor
            claim_info: Claim metadata (name, period, etc.)
        """
        self.results = processed_results
        self.claim_info = claim_info
        
    def generate_ct600l_summary(self) -> Dict:
        """
        Generate CT600L-ready summary data
        
        Returns:
            Dictionary containing CT600L summary information
        """
        summary = {
            'company_name': self.claim_info.get('company_name', ''),
            'period_start': self.claim_info.get('period_start', ''),
            'period_end': self.claim_info.get('period_end', ''),
            'total_rd_expenditure': float(self.results['total_qualifying_expenditure']),
            'staff_costs': float(self.results['staff_costs_with_nic']),
            'epw_costs': float(self.results['epw_costs']),
            'total_costs_claimed': float(self.results['qualifying_rd_costs']),
            'excluded_costs': float(self.results['excluded_costs']),
            'grant_adjustments': float(self.results.get('grant_adjustments', 0)),
            'number_of_line_items': len(self.results['line_items'])
        }
        
        return summary
    
    def generate_detailed_breakdown(self) -> pd.DataFrame:
        """
        Generate detailed breakdown of all line items
        
        Returns:
            Pandas DataFrame with detailed line item breakdown
        """
        line_items = []
        
        for item in self.results['line_items']:
            line_items.append({
                'Employee': item['employee_name'],
                'Description': item['description'],
                'Gross Cost': float(item['gross_cost']),
                'R&D Percentage': f"{item['rd_percentage']:.1%}",
                'Qualifying Cost': float(item['qualifying_cost']),
                'EPW': 'Yes' if item['is_epw'] else 'No',
                'Connected EPW': 'Yes' if item['epw_connected'] else 'No',
                'Excluded': 'Yes' if item['excluded'] else 'No',
                'Exclusion Reason': item['exclusion_reason'] or ''
            })
        
        return pd.DataFrame(line_items)
    
    def generate_employee_summary(self) -> pd.DataFrame:
        """
        Generate summary by employee
        
        Returns:
            Pandas DataFrame with employee summary
        """
        employee_summary = {}
        
        for item in self.results['line_items']:
            employee = item['employee_name']
            if employee not in employee_summary:
                employee_summary[employee] = {
                    'total_gross_cost': Decimal('0'),
                    'total_qualifying_cost': Decimal('0'),
                    'line_count': 0,
                    'is_epw': item['is_epw'],
                    'rd_percentage': item['rd_percentage']
                }
            
            employee_summary[employee]['total_gross_cost'] += item['gross_cost']
            employee_summary[employee]['total_qualifying_cost'] += item['qualifying_cost']
            employee_summary[employee]['line_count'] += 1
        
        # Convert to DataFrame
        summary_data = []
        for employee, data in employee_summary.items():
            summary_data.append({
                'Employee': employee,
                'Total Gross Cost': float(data['total_gross_cost']),
                'Total Qualifying Cost': float(data['total_qualifying_cost']),
                'R&D Percentage': f"{data['rd_percentage']:.1%}",
                'Line Items': data['line_count'],
                'EPW': 'Yes' if data['is_epw'] else 'No'
            })
        
        return pd.DataFrame(summary_data)
    
    def export_to_excel(self, include_detailed: bool = True) -> BytesIO:
        """
        Export all reports to Excel file
        
        Args:
            include_detailed: Whether to include detailed line item breakdown
            
        Returns:
            BytesIO object containing the Excel file
        """
        output = BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # CT600L Summary Sheet
            ct600l_summary = self.generate_ct600l_summary()
            summary_df = pd.DataFrame([ct600l_summary])
            summary_df.to_excel(writer, sheet_name='CT600L Summary', index=False)
            
            # Employee Summary Sheet
            employee_summary = self.generate_employee_summary()
            employee_summary.to_excel(writer, sheet_name='Employee Summary', index=False)
            
            # Detailed breakdown (if requested)
            if include_detailed:
                detailed_breakdown = self.generate_detailed_breakdown()
                detailed_breakdown.to_excel(writer, sheet_name='Detailed Breakdown', index=False)
            
            # Format the workbook
            self._format_excel_workbook(writer.book)
        
        output.seek(0)
        return output
    
    def _format_excel_workbook(self, workbook: openpyxl.Workbook):
        """
        Apply formatting to the Excel workbook
        
        Args:
            workbook: Openpyxl workbook object
        """
        # Define styles
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        center_alignment = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Format each worksheet
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            
            # Format headers
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = thin_border
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Apply borders to all cells with data
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        cell.border = thin_border
    
    def generate_narrative_template(self) -> Dict:
        """
        Generate narrative template for R&D claim documentation
        
        Returns:
            Dictionary containing narrative template sections
        """
        template = {
            'company_overview': {
                'title': 'Company Overview',
                'content': f"This R&D claim relates to {self.claim_info.get('company_name', '[Company Name]')} "
                          f"for the period {self.claim_info.get('period_start', '[Start Date]')} to "
                          f"{self.claim_info.get('period_end', '[End Date]')}.",
                'questions': [
                    'What is the nature of the company\'s business?',
                    'What R&D activities were undertaken during this period?',
                    'How many employees were involved in R&D activities?'
                ]
            },
            'rd_activities': {
                'title': 'R&D Activities',
                'content': 'The following R&D activities were undertaken during the claim period:',
                'questions': [
                    'What specific R&D projects were undertaken?',
                    'What technological advances were being sought?',
                    'What uncertainties existed that required resolution?',
                    'What was the baseline of existing knowledge?'
                ]
            },
            'staff_involvement': {
                'title': 'Staff Involvement',
                'content': f"A total of {len(set(item['employee_name'] for item in self.results['line_items']))} "
                          f"employees were involved in R&D activities during this period.",
                'questions': [
                    'Which employees were directly involved in R&D?',
                    'What was each employee\'s role in the R&D activities?',
                    'What percentage of their time was spent on R&D?',
                    'How was the R&D percentage determined for each employee?'
                ]
            },
            'epw_arrangements': {
                'title': 'Externally Provided Workers',
                'content': f"EPW costs totalling £{self.results['epw_costs']:,.2f} were incurred during this period.",
                'questions': [
                    'What EPW arrangements were in place?',
                    'Were the EPW arrangements connected or unconnected?',
                    'What R&D activities did the EPWs undertake?',
                    'How was the 65% cap applied to EPW costs?'
                ]
            },
            'excluded_costs': {
                'title': 'Excluded Costs',
                'content': f"Costs totalling £{self.results['excluded_costs']:,.2f} were excluded from the claim.",
                'questions': [
                    'What costs were excluded and why?',
                    'How were PILON and bonus payments treated?',
                    'Were there any other non-qualifying costs?'
                ]
            },
            'calculation_methodology': {
                'title': 'Calculation Methodology',
                'content': 'The R&D calculation was performed using the following methodology:',
                'questions': [
                    'How were R&D percentages determined?',
                    'What NIC uplift rate was applied?',
                    'How were grants and subsidies treated?',
                    'What quality assurance checks were performed?'
                ]
            }
        }
        
        return template
    
    def generate_audit_report(self, audit_trail: List[Dict]) -> Dict:
        """
        Generate audit report from audit trail
        
        Args:
            audit_trail: List of audit trail entries
            
        Returns:
            Dictionary containing audit report
        """
        report = {
            'claim_summary': {
                'total_line_items': len(self.results['line_items']),
                'total_qualifying_expenditure': float(self.results['total_qualifying_expenditure']),
                'processing_timestamp': datetime.now().isoformat(),
                'claim_reference': self.claim_info.get('reference', 'N/A')
            },
            'processing_summary': {
                'items_processed': len([entry for entry in audit_trail if entry['action'] == 'line_item_processed']),
                'items_excluded': len([entry for entry in audit_trail if entry['action'] == 'line_item_processed' and entry.get('excluded', False)]),
                'calculation_steps': len(audit_trail)
            },
            'exclusions': [
                {
                    'employee': entry['employee'],
                    'reason': entry['exclusion_reason'],
                    'gross_cost': entry['gross_cost']
                }
                for entry in audit_trail
                if entry['action'] == 'line_item_processed' and entry.get('excluded', False)
            ],
            'audit_trail': audit_trail
        }
        
        return report 